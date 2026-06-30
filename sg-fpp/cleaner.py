#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EE Listing 数据清洗自动化脚本（重构版）

与 Trial_7.py 逻辑完全一致，仅做以下调整以适配 Web 调用：
1. 移除 if __name__ == "__main__" 入口（由 app.py 调用）
2. 移除硬编码的文件路径配置（由调用方传入）
3. 错误处理改为 raise 异常，不再调用 sys.exit/input
4. 移除 input() 暂停（CLI 专用）
"""

import copy
import os
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 列名配置（按列名匹配，不依赖列位置）—— 与原脚本一致
# ============================================================

EE_COL_MAP = {
    "employee_name": "Employee Name",
    "login_code": "Login Code",
    "wecom_id": "WeCom ID",
    "identity_no": "Identity No.(last 4 digit)",
    "mobile_phone": "Mobile Phone No.",
    "start_date": "Eligibility Start Date",
    "end_date": "Eligibility End Date",
    "remarks": "Remarks",
}

ACTIVE_COL_MAP = {
    "worker": "Worker",
    "legal_name": "Full Legal Name",
    "wecom_id": "WeCom ID",
    "email": "Email - Primary Work",
    "probation_end_date": "Probation End Date",
    "national_id": "National ID",
    "phone": "Home Contact - Phone",
    "employee_type": "Employee Type",
}

TERM_COL_MAP = {
    "wecom_id": "WeCom ID",
    "last_day_of_work": "Last Day of Work",
    "employee_type": "Employee Type",
}

REMARKS_ADD_MEMBER = "Add Member"
REMARKS_TERMINATED = "Terminated"
REMARKS_CONFLICT = "冲突active+terminated"
REMARKS_LACK_PROBATION = "lack probation date"
REMARKS_TRANSFER_OUT = "国际调出员工"

FILL_ADD_MEMBER = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_TERMINATED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_CONFLICT = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_LACK_PROBATION = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

SORT_ORDER = {
    REMARKS_ADD_MEMBER: 1,
    REMARKS_CONFLICT: 2,
    REMARKS_LACK_PROBATION: 3,
    REMARKS_TRANSFER_OUT: 4,
    REMARKS_TERMINATED: 5,
    "": 6,
}


# ============================================================
# 列名匹配
# ============================================================

def find_column_index(ws, target_name, row=1):
    target_clean = target_name.lower().replace("\n", " ").replace("\r", " ").strip()

    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=row, column=col).value
        if header_val is None:
            continue
        header_clean = str(header_val).lower().replace("\n", " ").replace("\r", " ").strip()
        if header_clean == target_clean:
            return col

    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=row, column=col).value
        if header_val is None:
            continue
        header_clean = str(header_val).lower().replace("\n", " ").replace("\r", " ").strip()
        if target_clean in header_clean or header_clean in target_clean:
            return col

    return None


def build_col_index_map(ws, col_name_map):
    result = {}
    for logical_name, col_name in col_name_map.items():
        col_idx = find_column_index(ws, col_name)
        if col_idx is not None:
            result[logical_name] = col_idx
        else:
            print(f"  ⚠ 未找到列 '{col_name}'，相关功能将跳过")
    return result


# ============================================================
# 工具函数
# ============================================================

def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def safe_right(val, n):
    s = safe_str(val)
    return s[-n:] if len(s) >= n else s


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
        return None
    return None


def format_date(dt):
    if dt is None:
        return None
    return dt.strftime("%d-%b-%y")


def clear_row_fill(ws, row, max_col):
    no_fill = PatternFill(fill_type=None)
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = no_fill


def get_cell(ws, row, col_idx):
    if col_idx is None:
        return ""
    return ws.cell(row=row, column=col_idx).value


# ============================================================
# 主流程（与原 Trial_7.py 一致）
# ============================================================

def clean_ee_listing(ee_path, active_path, terminated_path, output_dir=None, output_filename=None):
    """EE Listing 数据清洗主流程（Web 版）

    Args:
        ee_path: 上月 EE Listing 文件路径
        active_path: 当月 Active Employee Report 文件路径
        terminated_path: 当月 Terminated Report 文件路径
        output_dir: 输出目录，None 时与 ee_path 同目录
        output_filename: 输出文件名（如 "APR_eelisting.xlsx"），None 时用当前月兜底

    Returns:
        output_path: 处理后的 Excel 文件路径

    Raises:
        FileNotFoundError: 任一输入文件不存在
        ValueError: 输入文件缺少必要列
    """
    print("=" * 60)
    print("EE Listing 数据清洗自动化")
    print("=" * 60)

    # 0. 读取文件 & 匹配列名
    print("\n[0] 读取文件 & 匹配列名...")

    for path, name in [(ee_path, "EE Listing"), (active_path, "Active Report"), (terminated_path, "Terminated Report")]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"找不到{name}文件: {path}")

    wb_ee = openpyxl.load_workbook(ee_path)
    ws_ee = wb_ee.active

    wb_active = openpyxl.load_workbook(active_path, data_only=True)
    ws_active = wb_active.active

    wb_term = openpyxl.load_workbook(terminated_path, data_only=True)
    ws_term = wb_term.active

    print(f"  EE Listing: {ws_ee.max_row - 1} 行, {ws_ee.max_column} 列")
    print(f"  Active Report: {ws_active.max_row - 1} 行, {ws_active.max_column} 列")
    print(f"  Terminated Report: {ws_term.max_row - 1} 行, {ws_term.max_column} 列")

    print("\n  匹配 EE Listing 列名:")
    ee_cols = build_col_index_map(ws_ee, EE_COL_MAP)
    for name, idx in ee_cols.items():
        print(f"    {EE_COL_MAP[name]} → {get_column_letter(idx)}列 (第{idx}列)")

    print("\n  匹配 Active Report 列名:")
    active_cols = build_col_index_map(ws_active, ACTIVE_COL_MAP)
    for name, idx in active_cols.items():
        print(f"    {ACTIVE_COL_MAP[name]} → {get_column_letter(idx)}列 (第{idx}列)")

    print("\n  匹配 Terminated Report 列名:")
    term_cols = build_col_index_map(ws_term, TERM_COL_MAP)
    for name, idx in term_cols.items():
        print(f"    {TERM_COL_MAP[name]} → {get_column_letter(idx)}列 (第{idx}列)")

    # 校验必要列
    for col_name in ["wecom_id", "remarks"]:
        if col_name not in ee_cols:
            raise ValueError(f"EE Listing 缺少必要列: {EE_COL_MAP[col_name]}")
    for col_name in ["wecom_id"]:
        if col_name not in active_cols:
            raise ValueError(f"Active Report 缺少必要列: {ACTIVE_COL_MAP[col_name]}")
    for col_name in ["wecom_id"]:
        if col_name not in term_cols:
            raise ValueError(f"Terminated Report 缺少必要列: {TERM_COL_MAP[col_name]}")

    ee_col_employee = ee_cols.get("employee_name")
    ee_col_login = ee_cols.get("login_code")
    ee_col_wecom = ee_cols["wecom_id"]
    ee_col_identity = ee_cols.get("identity_no")
    ee_col_phone = ee_cols.get("mobile_phone")
    ee_col_start = ee_cols.get("start_date")
    ee_col_end = ee_cols.get("end_date")
    ee_col_remarks = ee_cols["remarks"]

    # 1. 删除上月已标记 Terminated 的员工
    print("\n[1] 删除上月 Terminated 员工...")

    remarks_dist = {}
    for row in range(2, ws_ee.max_row + 1):
        r = safe_str(get_cell(ws_ee, row, ee_col_remarks))
        remarks_dist[r] = remarks_dist.get(r, 0) + 1
    print(f"  当前 Remarks 分布: {remarks_dist}")

    rows_to_delete = []
    for row in range(2, ws_ee.max_row + 1):
        remarks = safe_str(get_cell(ws_ee, row, ee_col_remarks))
        if remarks.lower().replace(" ", "") == "terminated":
            wecom_id = safe_str(get_cell(ws_ee, row, ee_col_wecom))
            rows_to_delete.append((row, wecom_id))

    for row, wid in reversed(rows_to_delete):
        ws_ee.delete_rows(row)

    print(f"  删除 {len(rows_to_delete)} 行 Terminated 员工")

    # 2. 清除上月 Add Member 标记
    print("\n[2] 清除上月 Add Member 标记...")

    add_member_cleared = 0
    clear_max_col = max(ws_ee.max_column, 50)
    for row in range(2, ws_ee.max_row + 1):
        remarks = safe_str(get_cell(ws_ee, row, ee_col_remarks))
        if remarks.lower().replace(" ", "") == "addmember":
            clear_row_fill(ws_ee, row, clear_max_col)
            ws_ee.cell(row=row, column=ee_col_remarks).value = ""
            add_member_cleared += 1

    print(f"  清除 {add_member_cleared} 行 Add Member 标记")

    # 过滤 Active Report 中的 Intern
    print("\n[0] 过滤 Active Report 中的 Intern...")

    active_col_employee_type = active_cols.get("employee_type")
    intern_count = 0

    if active_col_employee_type is not None:
        intern_wids = set()
        for row in range(2, ws_active.max_row + 1):
            emp_type = safe_str(get_cell(ws_active, row, active_col_employee_type))
            if emp_type.lower() == "intern":
                wid = safe_str(get_cell(ws_active, row, active_cols["wecom_id"]))
                if wid:
                    intern_wids.add(wid)
                intern_count += 1
        print(f"  过滤 {intern_count} 行 Intern 员工（{len(intern_wids)} 个 WeCom ID）")
    else:
        print("  ⚠ 未找到 Employee Type 列，跳过 Intern 过滤")
        intern_wids = set()

    # 3. 比对 Active Report
    print("\n[3] 比对 Active Report（已过滤 Intern）...")

    active_col_wecom = active_cols["wecom_id"]
    active_col_legal_name = active_cols.get("legal_name")
    active_col_email = active_cols.get("email")
    active_col_probation = active_cols.get("probation_end_date")
    active_col_national_id = active_cols.get("national_id")
    active_col_phone = active_cols.get("phone")

    active_data = {}
    for row in range(2, ws_active.max_row + 1):
        wid = safe_str(get_cell(ws_active, row, active_col_wecom))
        if not wid:
            continue
        if wid in intern_wids:
            continue
        active_data[wid] = {
            "legal_name": safe_str(get_cell(ws_active, row, active_col_legal_name)) if active_col_legal_name else "",
            "email": safe_str(get_cell(ws_active, row, active_col_email)),
            "national_id": safe_str(get_cell(ws_active, row, active_col_national_id)),
            "phone": safe_str(get_cell(ws_active, row, active_col_phone)),
            "probation_end_date": parse_date(get_cell(ws_active, row, active_col_probation)),
        }

    ee_wecom_ids = set()
    for row in range(2, ws_ee.max_row + 1):
        wid = safe_str(get_cell(ws_ee, row, ee_col_wecom))
        if wid:
            ee_wecom_ids.add(wid)

    new_wecom_ids = set(active_data.keys()) - ee_wecom_ids

    active_in_ee = ee_wecom_ids & set(active_data.keys())
    cleaned_remarks = 0
    for row in range(2, ws_ee.max_row + 1):
        wid = safe_str(get_cell(ws_ee, row, ee_col_wecom))
        remarks = safe_str(get_cell(ws_ee, row, ee_col_remarks))
        if wid in active_in_ee and remarks == "Active":
            ws_ee.cell(row=row, column=ee_col_remarks).value = ""
            cleaned_remarks += 1

    print(f"  在职员工: {len(active_in_ee)} 人")
    if cleaned_remarks > 0:
        print(f"  清理在职员工残留 Remarks: {cleaned_remarks} 行")
    print(f"  新增员工: {len(new_wecom_ids)} 人")

    error_probation = 0
    for wid in new_wecom_ids:
        emp = active_data[wid]
        new_row = ws_ee.max_row + 1

        if ee_col_employee is not None:
            ws_ee.cell(row=new_row, column=ee_col_employee).value = emp["legal_name"]
        if ee_col_login is not None:
            ws_ee.cell(row=new_row, column=ee_col_login).value = emp["email"]
        ws_ee.cell(row=new_row, column=ee_col_wecom).value = wid
        if ee_col_identity is not None:
            ws_ee.cell(row=new_row, column=ee_col_identity).value = safe_right(emp["national_id"], 4)
        if ee_col_phone is not None:
            ws_ee.cell(row=new_row, column=ee_col_phone).value = safe_right(emp["phone"], 8)
        if ee_col_end is not None:
            ws_ee.cell(row=new_row, column=ee_col_end).value = None

        if emp["probation_end_date"] is not None:
            if ee_col_start is not None:
                ws_ee.cell(row=new_row, column=ee_col_start).value = format_date(emp["probation_end_date"])
            ws_ee.cell(row=new_row, column=ee_col_remarks).value = REMARKS_ADD_MEMBER
            for col in range(1, 51):
                ws_ee.cell(row=new_row, column=col).fill = FILL_ADD_MEMBER
        else:
            if ee_col_start is not None:
                ws_ee.cell(row=new_row, column=ee_col_start).value = None
            ws_ee.cell(row=new_row, column=ee_col_remarks).value = REMARKS_LACK_PROBATION
            for col in range(1, 51):
                ws_ee.cell(row=new_row, column=col).fill = FILL_LACK_PROBATION
            error_probation += 1

    if error_probation > 0:
        print(f"  ⚠ {error_probation} 人 Probation End Date 为空，已标记 lack probation date")

    # 4. 比对 Terminated Report
    print("\n[4] 比对 Terminated Report...")

    term_col_wecom = term_cols["wecom_id"]
    term_col_last_day = term_cols.get("last_day_of_work")
    term_col_employee_type = term_cols.get("employee_type")

    term_intern_count = 0
    term_intern_wids = set()
    if term_col_employee_type is not None:
        for row in range(2, ws_term.max_row + 1):
            emp_type = safe_str(get_cell(ws_term, row, term_col_employee_type))
            if emp_type.lower() == "intern":
                wid = safe_str(get_cell(ws_term, row, term_col_wecom))
                if wid:
                    term_intern_wids.add(wid)
                term_intern_count += 1
        print(f"  过滤 {term_intern_count} 行 Intern 员工（{len(term_intern_wids)} 个 WeCom ID）")
    else:
        print("  ⚠ 未找到 Employee Type 列，跳过 Terminated Intern 过滤")

    term_data = {}
    for row in range(2, ws_term.max_row + 1):
        wid = safe_str(get_cell(ws_term, row, term_col_wecom))
        if not wid:
            continue
        if wid in term_intern_wids:
            continue
        term_data[wid] = {
            "last_day_of_work": parse_date(get_cell(ws_term, row, term_col_last_day)),
        }

    terminated_count = 0
    for row in range(2, ws_ee.max_row + 1):
        wid = safe_str(get_cell(ws_ee, row, ee_col_wecom))
        remarks = safe_str(get_cell(ws_ee, row, ee_col_remarks))
        if wid in term_data and remarks in ("", "Active", "Add Member"):
            last_day = term_data[wid]["last_day_of_work"]
            if ee_col_end is not None:
                if last_day is not None:
                    end_date = last_day + timedelta(days=1)
                    ws_ee.cell(row=row, column=ee_col_end).value = format_date(end_date)
                else:
                    ws_ee.cell(row=row, column=ee_col_end).value = None
            ws_ee.cell(row=row, column=ee_col_remarks).value = REMARKS_TERMINATED
            for col in range(1, 51):
                ws_ee.cell(row=row, column=col).fill = FILL_TERMINATED
            terminated_count += 1

    print(f"  标记 {terminated_count} 人 Terminated")

    # 5. 处理冲突
    print("\n[5] 检查 *error（同时在 Active 和 Terminated Report）...")

    active_wids = set(active_data.keys())
    term_wids = set(term_data.keys())
    both_wids = active_wids & term_wids

    error_conflict = 0
    for row in range(2, ws_ee.max_row + 1):
        wid = safe_str(get_cell(ws_ee, row, ee_col_wecom))
        remarks = safe_str(get_cell(ws_ee, row, ee_col_remarks))
        if wid in both_wids and remarks == REMARKS_TERMINATED:
            ws_ee.cell(row=row, column=ee_col_remarks).value = REMARKS_CONFLICT
            for col in range(1, 51):
                ws_ee.cell(row=row, column=col).fill = FILL_CONFLICT
            error_conflict += 1

    print(f"  标记 {error_conflict} 人冲突 active+terminated")

    # 6. 识别国际调出员工
    print("\n[6] 识别国际调出员工...")

    transfer_out_count = 0
    for row in range(2, ws_ee.max_row + 1):
        wid = safe_str(get_cell(ws_ee, row, ee_col_wecom))
        remarks = safe_str(get_cell(ws_ee, row, ee_col_remarks))

        if wid and wid not in active_wids and wid not in term_wids and remarks in ("", "Active"):
            ws_ee.cell(row=row, column=ee_col_remarks).value = REMARKS_TRANSFER_OUT
            transfer_out_count += 1

    print(f"  标记 {transfer_out_count} 人国际调出员工")

    # 7. 排序 + 输出
    print("\n[7] 排序 & 输出文件...")

    total_cols = max(ws_ee.max_column, 50)

    data_rows = []
    for row in range(2, ws_ee.max_row + 1):
        row_data = []
        row_styles = []
        for col in range(1, total_cols + 1):
            cell = ws_ee.cell(row=row, column=col)
            row_data.append(cell.value)
            row_styles.append({
                "font": copy.copy(cell.font),
                "border": copy.copy(cell.border),
                "fill": copy.copy(cell.fill),
                "alignment": copy.copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy.copy(cell.protection),
            })
        remarks = safe_str(row_data[ee_col_remarks - 1]) if ee_col_remarks else ""
        sort_key = SORT_ORDER.get(remarks, 99)
        original_order = row - 2
        data_rows.append((sort_key, original_order, row_data, row_styles))

    data_rows.sort(key=lambda x: (x[0], x[1]))

    ws_ee.delete_rows(2, ws_ee.max_row)

    for idx, (sort_key, orig_order, row_data, row_styles) in enumerate(data_rows):
        row_num = idx + 2
        for col_idx in range(total_cols):
            cell = ws_ee.cell(row=row_num, column=col_idx + 1)
            cell.value = row_data[col_idx]
            style = row_styles[col_idx]
            cell.font = style["font"]
            cell.border = style["border"]
            cell.fill = style["fill"]
            cell.alignment = style["alignment"]
            cell.number_format = style["number_format"]
            cell.protection = style["protection"]

    # 输出文件名（优先用调用方传入的，否则用当前月兜底）
    if output_filename:
        # 兜底：调用方没传 .xlsx 后缀
        if not output_filename.lower().endswith(".xlsx"):
            output_filename = output_filename + ".xlsx"
    else:
        now = datetime.now()
        month_abbr = now.strftime("%b")
        year_2digit = now.strftime("%y")
        output_filename = f"Tencent EE Listing_{month_abbr} {year_2digit}.xlsx"

    if not output_dir:
        output_dir = os.path.dirname(os.path.abspath(ee_path))
    output_path = os.path.join(output_dir, output_filename)

    wb_ee.save(output_path)

    pure_terminated = max(0, terminated_count - error_conflict)

    print("\n" + "=" * 60)
    print("✅ 数据清洗完成！")
    print("=" * 60)
    print(f"  过滤 Intern(Active): {intern_count} 人")
    print(f"  过滤 Intern(Terminated): {term_intern_count} 人")
    print(f"  删除上月 Terminated: {len(rows_to_delete)} 人")
    print(f"  清除 Add Member 标记: {add_member_cleared} 人")
    print(f"  新增员工(Add Member): {len(new_wecom_ids) - error_probation - error_conflict} 人")
    print(f"  本月 Terminated: {pure_terminated} 人")
    print(f"  *error(Probation 为空): {error_probation} 人")
    print(f"  冲突 active+terminated: {error_conflict} 人")
    print(f"  国际调出员工: {transfer_out_count} 人")
    print(f"\n  输出文件: {output_path}")

    wb_ee.close()
    wb_active.close()
    wb_term.close()

    return output_path
