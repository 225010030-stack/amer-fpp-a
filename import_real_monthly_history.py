#!/usr/bin/env python3
"""Build monthly dashboard history from real xlsx data folders."""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


TARGET_MONTHS = {"202601", "202602", "202603", "202604", "202605"}
AMOUNT_RE = re.compile(r"\$?\s*([0-9][0-9,]*\.?[0-9]*)")


@dataclass
class InputRow:
    period: str
    country: str
    entity: str
    vendor: str
    fee_type: str
    amount: float
    currency: str
    center: str


def parse_month_from_path(path: Path) -> str:
    for p in path.parts:
        if p in TARGET_MONTHS:
            return p
        m = re.search(r"(20\d{4})", p)
        if m and m.group(1) in TARGET_MONTHS:
            return m.group(1)
    m2 = re.search(r"(20\d{4})", path.name)
    if m2 and m2.group(1) in TARGET_MONTHS:
        return m2.group(1)
    return ""


def parse_vendor_from_path(path: Path) -> str:
    # e.g. "1. Anthem - Medical insurance" -> "Anthem"
    for p in path.parts:
        if " - " in p and any(ch.isdigit() for ch in p[:2]):
            part = p.split(" - ", 1)[0]
            return part.split(".", 1)[-1].strip()
    # fallback
    stem = path.stem
    if "_" in stem:
        return stem.split("_", 1)[1].split("-", 1)[0].strip()
    return "Vendor"


def parse_entity_from_file(path: Path) -> str:
    # e.g. "Tencent America LLC_ Anthem - 202601.xlsx"
    stem = path.stem
    if "_" in stem:
        return stem.split("_", 1)[0].strip()
    return "Entity"


def parse_country_from_root(path: Path) -> str:
    parts = {p.upper() for p in path.parts}
    if "CAN" in parts:
        return "CAN"
    return "US"


def extract_amount_from_text(text: str) -> Optional[float]:
    best: Optional[float] = None
    for m in AMOUNT_RE.finditer(text or ""):
        raw = m.group(1).replace(",", "")
        try:
            val = float(raw)
        except ValueError:
            continue
        if val <= 0:
            continue
        if best is None or val > best:
            best = val
    return best


def read_hc_rows(xlsx_path: Path) -> list[InputRow]:
    month = parse_month_from_path(xlsx_path)
    if not month:
        return []
    country = parse_country_from_root(xlsx_path)
    vendor = parse_vendor_from_path(xlsx_path)
    entity = parse_entity_from_file(xlsx_path)
    currency = "CAD" if country == "CAN" else "USD"

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return []
    sheet = wb["核算明细信息"] if "核算明细信息" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    col = {name: idx for idx, name in enumerate(header)}
    amount_idx = col.get("核算金额")
    fee_idx = col.get("费用类型（物料）")
    center_idx = col.get("成本中心编码")
    if amount_idx is None:
        return []

    out: list[InputRow] = []
    for r in sheet.iter_rows(min_row=2, values_only=True):
        raw_amount = r[amount_idx]
        try:
            amount = float(raw_amount or 0)
        except (TypeError, ValueError):
            continue
        if amount == 0:
            continue
        fee = str(r[fee_idx]).strip() if fee_idx is not None and r[fee_idx] is not None else "Employee Cost"
        center = str(r[center_idx]).strip() if center_idx is not None and r[center_idx] is not None else ""
        out.append(
            InputRow(
                period=month,
                country=country,
                entity=entity,
                vendor=vendor,
                fee_type=fee,
                amount=amount,
                currency=currency,
                center=center,
            )
        )
    return out


def read_optum_rows(path: Path) -> list[InputRow]:
    # Optum files under invoice folders often store amount in filename.
    month = parse_month_from_path(path)
    if not month:
        return []
    amount = extract_amount_from_text(path.name)
    if amount is None:
        return []
    return [
        InputRow(
            period=month,
            country="US",
            entity="Tencent America LLC",
            vendor="Optum",
            fee_type="FSA Weekly Claim",
            amount=amount,
            currency="USD",
            center="",
        )
    ]


def read_legalshield_rows(pdf_path: Path) -> list[InputRow]:
    month = parse_month_from_path(pdf_path)
    if not month:
        return []
    # First try filename.
    amount = extract_amount_from_text(pdf_path.name)
    # Then try PDF text if needed.
    if amount is None:
        try:
            from pypdf import PdfReader  # type: ignore

            reader = PdfReader(str(pdf_path))
            txt = "\n".join((p.extract_text() or "") for p in reader.pages[:2])
            amount = extract_amount_from_text(txt)
        except Exception:
            amount = None
    if amount is None:
        return []
    return [
        InputRow(
            period=month,
            country="US",
            entity="Tencent America LLC",
            vendor="LegalShield",
            fee_type="ID Legal Insurance",
            amount=amount,
            currency="USD",
            center="",
        )
    ]


def write_prefill_csv(path: Path, rows: list[InputRow]) -> None:
    headers = [
        "期间",
        "国家",
        "主体",
        "供应商",
        "费用类型",
        "FPP场景",
        "Document Category",
        "交易币种",
        "支付币种",
        "Payment Description",
        "Invoice策略",
        "合同号",
        "账单金额(不含税)",
        "税额",
        "含税总额",
        "Requester",
        "Business Reviewer",
        "Department",
        "Center",
        "成本中心校验状态",
        "期间一致性状态",
        "金额校验状态",
        "可提交(Y/N)",
        "阻断原因",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "期间": r.period,
                    "国家": r.country,
                    "主体": r.entity,
                    "供应商": r.vendor,
                    "费用类型": r.fee_type,
                    "FPP场景": "Benefits Insurance",
                    "Document Category": "Invoice",
                    "交易币种": r.currency,
                    "支付币种": r.currency,
                    "Payment Description": f"{r.vendor} {r.fee_type}",
                    "Invoice策略": "WITH INVOICE",
                    "合同号": "",
                    "账单金额(不含税)": f"{r.amount:.2f}",
                    "税额": "0",
                    "含税总额": f"{r.amount:.2f}",
                    "Requester": "",
                    "Business Reviewer": "",
                    "Department": "",
                    "Center": r.center,
                    "成本中心校验状态": "ok",
                    "期间一致性状态": "ok",
                    "金额校验状态": "ok",
                    "可提交(Y/N)": "Y",
                    "阻断原因": "",
                }
            )


def call_stats(api_base: str, root: Path, region: str, csv_path: Path) -> Optional[dict]:
    cmd = [
        "curl",
        "-s",
        "-X",
        "POST",
        f"{api_base}/api/upload-and-generate",
        "-F",
        "action=stats_analysis_upload",
        "-F",
        f"region={region}",
        "-F",
        f"root={root}",
        "-F",
        f"file=@{csv_path}",
    ]
    out = subprocess.check_output(cmd, text=True)
    data = json.loads(out)
    return data.get("stats_summary")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import real monthly data and build dashboard history")
    parser.add_argument("--source-root", default="/Users/zhangwenjing/Desktop/工作提效_副本", help="Real data source root")
    parser.add_argument("--workspace-root", default="/Users/zhangwenjing/Desktop/工作提效", help="Workspace root")
    parser.add_argument("--api-base", default="http://127.0.0.1:18082", help="Backend API")
    args = parser.parse_args()

    source_root = Path(args.source_root).resolve()
    workspace_root = Path(args.workspace_root).resolve()
    ingest_dir = workspace_root / "上传输入" / "月度导入真实_202601_202605"
    ingest_dir.mkdir(parents=True, exist_ok=True)

    all_xlsx = list(source_root.rglob("*.xlsx"))
    all_pdf = list(source_root.rglob("*.pdf"))
    month_rows: dict[tuple[str, str], list[InputRow]] = {}
    coverage: dict[str, dict[str, int]] = {"US": {}, "CAN": {}}
    for x in all_xlsx:
        month = parse_month_from_path(x)
        if not month or month not in TARGET_MONTHS:
            continue
        if "optum" in str(x).lower():
            rows = read_optum_rows(x)
        else:
            rows = read_hc_rows(x)
        for r in rows:
            key = (r.country, r.period)
            month_rows.setdefault(key, []).append(r)
            vkey = f"{r.vendor}|{r.fee_type}"
            coverage.setdefault(r.country, {})
            coverage[r.country][vkey] = coverage[r.country].get(vkey, 0) + 1

    for p in all_pdf:
        month = parse_month_from_path(p)
        if not month or month not in TARGET_MONTHS:
            continue
        if "legalshield" not in str(p).lower():
            continue
        rows = read_legalshield_rows(p)
        for r in rows:
            key = (r.country, r.period)
            month_rows.setdefault(key, []).append(r)
            vkey = f"{r.vendor}|{r.fee_type}"
            coverage.setdefault(r.country, {})
            coverage[r.country][vkey] = coverage[r.country].get(vkey, 0) + 1

    # Write prefill CSVs per region-month
    csv_paths: dict[tuple[str, str], Path] = {}
    for (region, month), rows in sorted(month_rows.items()):
        out = ingest_dir / f"{region}_prefill_{month}.csv"
        write_prefill_csv(out, rows)
        csv_paths[(region, month)] = out

    history: dict[str, dict[str, dict]] = {"US": {}, "CAN": {}}
    for (region, month), p in sorted(csv_paths.items()):
        summary = call_stats(args.api_base, workspace_root, region, p)
        if summary:
            rset = {
                (r.vendor.strip(), r.fee_type.strip())
                for r in month_rows.get((region, month), [])
                if r.vendor.strip() or r.fee_type.strip()
            }
            summary["coverage_type_count"] = len(rset)
            summary["coverage_types"] = sorted([f"{a}|{b}" for a, b in rset])
            history[region][month] = summary

    out_us = workspace_root / "上传输出" / "月度看板历史_US_真实数据_202601_202605.json"
    out_can = workspace_root / "上传输出" / "月度看板历史_CAN_真实数据_202601_202605.json"
    coverage_file = workspace_root / "上传输出" / "月度看板覆盖说明_真实数据_202601_202605.json"
    out_us.write_text(json.dumps(history["US"], ensure_ascii=False, indent=2), encoding="utf-8")
    out_can.write_text(json.dumps(history["CAN"], ensure_ascii=False, indent=2), encoding="utf-8")
    coverage_file.write_text(json.dumps(coverage, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"DONE: {out_us}")
    print(f"DONE: {out_can}")
    print(f"DONE: {coverage_file}")
    print(f"DONE: {ingest_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
